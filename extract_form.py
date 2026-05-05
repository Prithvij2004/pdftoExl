"""
PDF -> Excel form extractor (single-file CLI).

Usage:
    export LLAMA_CLOUD_API_KEY=llx-...
    python extract_form.py path/to/form.pdf -o out.xlsx

Pipeline: LlamaExtract (per_doc, agentic) with a Pydantic schema modeled on
docs/requirements.md -> openpyxl writer producing the 7-column workbook.
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from typing import List, Literal, Optional

from pydantic import BaseModel, Field

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    def _load_env_file(path: Path) -> None:
        if not path.exists():
            return
        for raw in path.read_text().splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = val
    _load_env_file(Path(__file__).parent / ".env")

# ----------------------- Schema (drives LlamaExtract) -----------------------

QuestionType = Literal[
    "Display",
    "Text Box",
    "Text Area",
    "Date",
    "Number",
    "Signature",
    "Radio Button",
    "Checkbox",
    "Checkbox Group",
    "Dropdown",
    "Group Table",
]


class FormRow(BaseModel):
    """One output row in the final Excel workbook.

    The model must produce rows in document order. Sequence is assigned later.
    """

    section: Optional[str] = Field(
        default=None,
        description=(
            "The section heading this row belongs to. As soon as a new "
            "section heading appears in the document, EVERY subsequent row "
            "(starting from the section-marker row itself) must carry the "
            "NEW heading text in this field — never the previous section's "
            "name. The value stays constant until the next section heading "
            "appears, at which point it switches again. Rows that appear "
            "before the very first section heading have a blank `section`."
        ),
    )
    question_rule: Optional[str] = Field(
        default=None,
        description=(
            "Optional rule/prelude that scopes the question. Populate from "
            "leading 'For <population>:' prefixes or italic prelude blocks "
            "that immediately precede the question. Otherwise leave blank."
        ),
    )
    question_type: QuestionType = Field(
        description=(
            "The canonical type of this row. MUST be exactly one of: "
            "Display, Text Box, Text Area, Date, Number, Signature, "
            "Radio Button, Checkbox, Checkbox Group, Dropdown, Group Table.\n"
            "Rules:\n"
            "- Display: any instruction, knowledge, narrative, header note, "
            "  or guidance block that does NOT ask the form-filler for an "
            "  answer. For Display rows the raw text goes into question_text "
            "  VERBATIM and answer_text MUST stay blank.\n"
            "- Text Box: single-line free-text input. Any 'Label: ____' blank "
            "  whose answer fits on one line is Text Box. This INCLUDES "
            "  'Description of documentation attached: ____', 'Name of CBRA "
            "  facility: ____', 'Specify: ____', 'Acute event: ____', etc. "
            "  The number of underscores is NOT a signal — many forms render "
            "  one logical line with a long underline.\n"
            "- Text Area: ONLY when the form clearly reserves MULTIPLE "
            "  consecutive blank lines (e.g. 4+ stacked underline rows) for "
            "  the answer, OR the prompt explicitly says 'describe in detail' "
            "  / 'provide a detailed description' / 'explain' AND the layout "
            "  shows a multi-line answer area. If unsure, choose Text Box.\n"
            "- Date: any field labeled date/dob/birth or formatted MM/DD/YYYY.\n"
            "- Number: fields labeled score/count/age/number/# or numeric-only.\n"
            "- Signature: a signature line (lone 'signature' label).\n"
            "- Radio Button: a single prompt followed by mutually-exclusive "
            "  options where exactly one must be selected. The prompt becomes "
            "  question_text and the options become answer_text (one option "
            "  per line, verbatim, in document order). DO NOT split this into "
            "  a Display row for the prompt plus a separate row for the "
            "  options — they are ONE Radio Button row. Residence/marital/"
            "  relationship/location-style option lists are inherently "
            "  mutually exclusive and are Radio Button (or Dropdown), never "
            "  Checkbox Group, even when each option ends with '—specify…'.\n"
            "- Checkbox: a SINGLE standalone check item with no sibling "
            "  options (a lone yes/acknowledgement-style box).\n"
            "- Checkbox Group: ONLY when ALL of the following are true: "
            "  (a) the prompt explicitly says 'check all that apply' / "
            "  'select all that apply' OR the options are semantically "
            "  independent flags; AND (b) the options are short, parallel, "
            "  self-contained statements (NOT multi-sentence criterion "
            "  paragraphs each with their own follow-up questions); AND "
            "  (c) NONE of the options have indented 'o'-bulleted child "
            "  questions or follow-up Text Area/Text Box prompts beneath "
            "  them. If any option has its OWN gated follow-up row, the "
            "  parent is NOT a Checkbox Group — each option is its own "
            "  standalone Checkbox row with the follow-up as a sibling "
            "  gated row. Default AWAY from Checkbox Group when in doubt — "
            "  prefer Radio Button or a series of standalone Checkbox rows.\n"
            "- Dropdown: a select-one field with a short (<=7) mutually-"
            "  exclusive option list rendered inline (e.g. 'AM / PM').\n"
            "- Group Table: a multi-row/multi-column table. The FIRST row "
            "  has question_type='Group Table' and question_text = the "
            "  table's title verbatim. If the table has no visible title, "
            "  generate a short descriptive title from the columns (e.g. "
            "  'Hospital admissions table'). Then emit ONE row per column "
            "  header, in left-to-right order, where question_text is the "
            "  bare column label EXACTLY as printed (do NOT prepend the "
            "  table title or any prefix) and question_type is inferred "
            "  from the column label (date|admit date|discharge date -> "
            "  Date; reason|description|notes -> Text Box; etc.). Drop "
            "  data rows entirely. EACH distinct table is its OWN Group "
            "  Table parent + its OWN column rows — never merge two "
            "  separate tables (different titles or different column sets) "
            "  into a single Group Table.\n"
            "Composite 'X with Text Area' types are NOT allowed.\n"
            "GLYPH-INDEPENDENCE: Tick (✓), check (☑), square (☐), or bullet "
            "  glyphs are decorative — they do NOT determine the type. Decide "
            "  purely from semantics."
        ),
    )
    question_text: str = Field(
        description=(
            "The prompt text shown to the form-filler. For Group Table this "
            "is the table title. For section-marker rows (see below) this is "
            "the literal string 'New Section'."
        ),
    )
    branching_logic: Optional[str] = Field(
        default=None,
        description=(
            "Branching gate for child rows. ONLY two templates are allowed:\n"
            "1. 'If Q{n} = checked(selected)' — when this row is gated by a "
            "   checkbox parent.\n"
            "2. 'Display if Q{n} = \"{option_text_verbatim}\"' — when this row "
            "   is gated by a specific radio/dropdown option of a parent. "
            "   Option text MUST be quoted exactly as it appears in the "
            "   parent's answer_text.\n"
            "Use Q{n} where n is the 1-based sequence number of the parent "
            "row in this output. Leave blank for ungated rows. Indent-based "
            "gating is inferred only within a single page; never across pages. "
            "Free-text expressions are forbidden."
        ),
    )
    answer_text: Optional[str] = Field(
        default=None,
        description=(
            "Row payload depending on question_type:\n"
            "- Radio Button / Dropdown / Checkbox Group: the option list, "
            "  one option per line, verbatim, in document order.\n"
            "- Section marker rows: the section heading text.\n"
            "- Display: MUST be blank. The instruction/knowledge text "
            "  belongs in question_text verbatim; never split prose between "
            "  question_text and answer_text.\n"
            "- Text Box / Text Area / Date / Number / Signature / Checkbox / "
            "  Group Table: leave blank.\n"
        ),
    )


class ExtractedForm(BaseModel):
    """Top-level extraction result.

    Emit rows in strict document order. The post-processor will assign dense
    1..N Sequence values.
    """

    rows: List[FormRow] = Field(
        description=(
            "All rows of the output workbook, in document order.\n\n"
            "Required behaviors:\n"
            "1. SECTIONS. When a section heading appears, emit a MARKER row "
            "   with question_type='Display', question_text='New Section', "
            "   answer_text=<heading text>. Every row from the marker onward "
            "   carries the NEW heading in `section` until the next heading. "
            "   Rows before the first heading have blank `section`. Emit the "
            "   marker once per boundary. Free-standing prose between a "
            "   heading and the first question is its OWN Display row.\n"
            "2. BLANK FIELDS. 'Label: ____' blanks become rows typed by "
            "   label: date|dob|birth -> Date; score|count|age|number|# -> "
            "   Number; lone 'signature' -> Signature; else Text Box. Long "
            "   underlines do NOT make a field Text Area — only stacked "
            "   multi-line answer areas do.\n"
            "3. OPTIONS — MERGE INTO ONE ROW. A prompt followed by its "
            "   option list is ONE row, NOT two. Put the prompt in "
            "   question_text and the options in answer_text (one per line). "
            "   NEVER emit a Display row for the prompt and a separate row "
            "   for the options — that is wrong. Type selection:\n"
            "   - Inline short list like 'AM / PM' -> Dropdown.\n"
            "   - Mutually-exclusive vertical list (residence, location, "
            "     relationship, status, yes/no) -> Radio Button. This is the "
            "     default for any vertical option list with square/bullet "
            "     glyphs unless the prompt explicitly says 'check all that "
            "     apply' / 'select all that apply'.\n"
            "   - 'check all that apply' / 'select all that apply' / "
            "     independent flags -> Checkbox Group.\n"
            "   - A single standalone box with no siblings -> Checkbox.\n"
            "   Glyph shape (☐ ☑ ✓ •) is NOT a signal — semantics decide.\n"
            "4. TABLES. EACH distinct table becomes ITS OWN Group Table "
            "   row (question_text = table title verbatim, or a short "
            "   generated title if none exists) followed by ONE row per "
            "   column header. Column-header rows: question_text = the bare "
            "   column label (no table-title prefix), question_type = "
            "   inferred from the label (Date for date/admit/discharge; "
            "   Text Box for reason/description; Number for count/score). "
            "   Drop all data rows. NEVER merge two visually separate "
            "   tables (different titles, different columns, different "
            "   purposes) into one Group Table — emit each separately.\n"
            "5. CHECKBOX FOLLOW-UPS. An indented 'o'-bulleted prompt or "
            "   'Description of documentation attached: ____' line under a "
            "   parent checkbox is a SIBLING gated row, NOT an option of "
            "   the checkbox. Set branching_logic = "
            "   'If Q{n} = checked(selected)'.\n"
            "6. SPECIFIERS. A 'Specify' / 'Specify other' blank after an "
            "   option containing 'other'/'specify' is a SEPARATE Text Box "
            "   row gated on that exact option: branching_logic = "
            "   'Display if Q{n} = \"<option verbatim>\"'.\n"
            "7. REPEATING TEMPLATES — EMIT ONCE. Some forms repeat the SAME "
            "   field block multiple times (e.g. four identical 'Fall #' "
            "   blocks across pages 10-11; multiple 'Applicant Name / SSN / "
            "   DOB' identifier banners on every page; repeated table data "
            "   rows). Emit the field set EXACTLY ONCE — the FIRST "
            "   occurrence — and do NOT emit duplicates of the same "
            "   template, even when subsequent copies appear on later pages. "
            "   Two blocks count as 'the same template' when their labels "
            "   match (case-insensitive) in the same order.\n"
            "8. NOISE TO DROP ENTIRELY. Do NOT emit rows for: agency logos, "
            "   form codes ('TC0175 (Rev. 8-2-16)', 'RDA 2047'), page "
            "   numbers, recurring page-header titles ('Safety Determination "
            "   Request Form' repeated on every page), footers, watermarks, "
            "   or any decorative repeat band. The form TITLE itself is also "
            "   not a question — drop it. Start the output at the first "
            "   real field or prose block.\n"
            "9. CROSS-PAGE STITCHING. A question whose options or answer "
            "   area spill onto the next page merges into ONE row. This "
            "   applies to Radio Button, Checkbox Group, and Dropdown: if "
            "   options 1-2 are at the bottom of page N and option 3 is at "
            "   the top of page N+1 with no intervening section heading, "
            "   ALL options belong in the SAME parent row's answer_text. "
            "   Do NOT emit option 3 as a separate Checkbox row. Treat the "
            "   options as a continuous list across the page break.\n"
            "9a. NO FALSE GROUPING. Do NOT bundle a list of long criterion "
            "    paragraphs into a single Checkbox Group when each "
            "    criterion has its OWN indented follow-up question(s) "
            "    beneath it (e.g. 'Provide a detailed description...', "
            "    'Describe how often...', 'Document below...'). That "
            "    layout is a series of independent Checkbox parents, each "
            "    with sibling gated follow-up rows. The presence of any "
            "    'o'-bulleted child question under an option is "
            "    DEFINITIVE proof it is NOT a Checkbox Group option — it "
            "    is its own Checkbox parent.\n"
            "10. QUESTION RULE. Populate question_rule from leading "
            "    'For <population>:' prefixes or italic prelude blocks.\n"
            "11. NO EMPTY ROWS. Never emit a row with empty question_text. "
            "    If you have no prompt text, drop the row entirely.\n"
            "12. ORDER. Rows must follow document reading order exactly."
        ),
    )


# ----------------------- LlamaExtract call -----------------------

SYSTEM_PROMPT = (
    "You are converting a government intake/assessment/eligibility PDF form "
    "into a strict 7-column row schema for downstream re-rendering. The PDF "
    "is digitally generated (not scanned) and may or may not have AcroForm "
    "widgets. Recover the form's logical structure — sections, questions, "
    "fields, options, tables, gated follow-ups — purely from layout, "
    "indentation, and blank-field patterns. Glyph shape is decorative and "
    "must NEVER drive type choice.\n\n"
    "CRITICAL RULES (these are the most common failure modes):\n"
    "1. Drop ALL noise: agency logos, form codes (e.g. 'TC0175 (Rev. ...)'), "
    "   'RDA' codes, page numbers, recurring page-header titles, footers, "
    "   identifier banners that repeat on every page (e.g. 'Applicant Name "
    "   / SSN / DOB' shown on every page header). The repeated header "
    "   appears on the FIRST page only conceptually, but for this schema "
    "   the safe choice is to drop the page-header banner entirely. Real "
    "   identifier fields embedded in the form body are kept.\n"
    "2. Repeating field blocks (e.g. four identical 'Fall #' blocks; the "
    "   same medical condition row repeated; table data rows) are emitted "
    "   ONCE — the first instance only. Never duplicate.\n"
    "3. A prompt and its option list are ONE row, not two. Put the prompt "
    "   in question_text and options in answer_text. Do NOT emit a separate "
    "   Display row for the prompt followed by a Checkbox Group row for "
    "   the options — that is the wrong split.\n"
    "4. Vertical option lists rendered with bullet/square glyphs are Radio "
    "   Button by default. Only use Checkbox Group when the prompt clearly "
    "   says 'check all that apply' AND the options are short parallel "
    "   self-contained statements with NO indented follow-up questions. "
    "   If any option has its own 'o'-bulleted child prompt beneath it, "
    "   the option is its own standalone Checkbox parent — NOT a Checkbox "
    "   Group entry. A list of multi-sentence criterion paragraphs each "
    "   with their own follow-up Text Area is a series of Checkbox rows, "
    "   not a Checkbox Group.\n"
    "4a. Cross-page option stitching: if a Radio Button / Checkbox Group / "
    "    Dropdown's options span a page break (e.g. 2 options at the "
    "    bottom of page N, 1 option at the top of page N+1), MERGE them "
    "    into one row's answer_text. Never emit the spilled option as a "
    "    separate row.\n"
    "4b. Group Table: each distinct table is its OWN Group Table row "
    "    (question_text = table title verbatim, or a short generated "
    "    title if none exists) followed by ONE row per column header "
    "    (question_text = bare column label, NOT prefixed with the table "
    "    title). NEVER merge two separate tables into one Group Table.\n"
    "5. Text Area is ONLY for true multi-line answer areas (4+ stacked "
    "   underline rows or explicit 'describe in detail'). Single-line "
    "   labeled blanks like 'Description of documentation attached: ____' "
    "   are Text Box even with long underlines.\n"
    "6. Never emit a row with empty question_text.\n\n"
    "Follow the schema field descriptions exactly. The question_type "
    "taxonomy is closed; branching_logic has only two templates. Emit "
    "every row in document reading order."
)


def run_extract(pdf_path: Path, api_key: str) -> ExtractedForm:
    from llama_cloud import LlamaCloud

    client = LlamaCloud(api_key=api_key)

    print(f"[1/3] Uploading {pdf_path.name}...", file=sys.stderr)
    file_obj = client.files.create(file=str(pdf_path), purpose="extract")

    print("[2/3] Submitting extraction job...", file=sys.stderr)
    job = client.extract.create(
        file_input=file_obj.id,
        configuration={
            "data_schema": ExtractedForm.model_json_schema(),
            "extraction_target": "per_doc",
            "tier": "agentic",
            "system_prompt": SYSTEM_PROMPT,
            "cite_sources": True,
            "confidence_scores": True,
        },
    )

    print(f"[3/3] Polling job {job.id}...", file=sys.stderr)
    while job.status not in ("COMPLETED", "FAILED", "CANCELLED"):
        time.sleep(2)
        job = client.extract.get(job.id)
        print(f"      status={job.status}", file=sys.stderr)

    if job.status != "COMPLETED":
        raise RuntimeError(f"Extraction job ended with status={job.status}")

    return ExtractedForm.model_validate(job.extract_result)


# ----------------------- Post-processing safety net -----------------------

import re

_NOISE_PATTERNS = [
    re.compile(r"^TC\d{3,5}\s*\(Rev", re.I),
    re.compile(r"^RDA\s*\d+", re.I),
    re.compile(r"^page\s*\d+\s*(of\s*\d+)?$", re.I),
    re.compile(r"^\d+\s+Safety Determination Request Form\s*$", re.I),
    re.compile(r"logo$", re.I),
]

_LABELED_BLANK_HINTS = (
    "description of documentation attached",
    "name of",
    "specify",
    "acute event",
    "treatment required",
    "duration of",
    "credentials",
    "printed name",
    "reason for",
)


def _is_noise(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return True
    return any(p.search(t) for p in _NOISE_PATTERNS)


def _template_key(rows: List[FormRow], start: int, length: int) -> tuple:
    return tuple(
        (r.question_type, (r.question_text or "").strip().lower())
        for r in rows[start : start + length]
    )


def _dedupe_repeating_templates(rows: List[FormRow]) -> List[FormRow]:
    """Detect runs of identical field blocks (e.g. four 'Fall #' blocks) and
    keep only the first occurrence."""
    if not rows:
        return rows
    # Find anchor labels that look like template starts (short, ends with '#'
    # or matches 'fall #', 'block', etc.). Generic: any row whose question_text
    # appears 2+ times anywhere in the list is a candidate anchor.
    text_counts: dict[str, int] = {}
    for r in rows:
        key = (r.question_text or "").strip().lower()
        if key:
            text_counts[key] = text_counts.get(key, 0) + 1

    repeat_anchors = {k for k, c in text_counts.items() if c >= 2}
    if not repeat_anchors:
        return rows

    seen_blocks: set[tuple] = set()
    out: List[FormRow] = []
    i = 0
    while i < len(rows):
        key = (rows[i].question_text or "").strip().lower()
        if key in repeat_anchors:
            # measure block length: extend until next anchor or end
            j = i + 1
            while j < len(rows):
                jk = (rows[j].question_text or "").strip().lower()
                if jk in repeat_anchors and jk == key:
                    break
                j += 1
            block_sig = _template_key(rows, i, j - i)
            if block_sig in seen_blocks:
                i = j
                continue
            seen_blocks.add(block_sig)
            out.extend(rows[i:j])
            i = j
        else:
            out.append(rows[i])
            i += 1
    return out


def _coerce_text_area_to_text_box(row: FormRow) -> FormRow:
    if row.question_type != "Text Area":
        return row
    qt = (row.question_text or "").strip().lower().rstrip(":")
    if any(h in qt for h in _LABELED_BLANK_HINTS) and len(qt) < 80:
        row.question_type = "Text Box"
    return row


def post_process(form: ExtractedForm) -> ExtractedForm:
    cleaned: List[FormRow] = []
    for r in form.rows:
        if _is_noise(r.question_text or ""):
            continue
        if not (r.question_text or "").strip():
            continue
        cleaned.append(_coerce_text_area_to_text_box(r))
    cleaned = _dedupe_repeating_templates(cleaned)
    form.rows = cleaned
    return form


# ----------------------- Excel writer -----------------------

HEADERS = [
    "Section",
    "Sequence",
    "Question Rule",
    "Question Type",
    "Question Text",
    "Branching Logic",
    "Answer Text",
]


def write_xlsx(form: ExtractedForm, out_path: Path) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Form"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    widths = [22, 10, 24, 16, 60, 40, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    for seq, row in enumerate(form.rows, start=1):
        ws.append([
            row.section or "",
            seq,
            row.question_rule or "",
            row.question_type,
            row.question_text or "",
            row.branching_logic or "",
            row.answer_text or "",
        ])
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=seq + 1, column=col_idx).alignment = body_align

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(form.rows)


# ----------------------- CLI -----------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Convert a PDF form to the 7-column .xlsx schema using LlamaExtract.")
    parser.add_argument("pdf", type=Path, help="Path to the input PDF.")
    parser.add_argument("-o", "--output", type=Path, default=None, help="Output .xlsx path (default: alongside the PDF).")
    parser.add_argument("--dump-json", type=Path, default=None, help="Also write the raw extraction JSON here for inspection.")
    args = parser.parse_args()

    if not args.pdf.exists():
        print(f"error: {args.pdf} not found", file=sys.stderr)
        return 2

    api_key = os.environ.get("LLAMA_CLOUD_API_KEY")
    if not api_key:
        print("error: LLAMA_CLOUD_API_KEY is not set", file=sys.stderr)
        return 2

    out_path = args.output or args.pdf.with_suffix(".xlsx")

    form = run_extract(args.pdf, api_key)
    form = post_process(form)

    if args.dump_json:
        args.dump_json.parent.mkdir(parents=True, exist_ok=True)
        args.dump_json.write_text(form.model_dump_json(indent=2))

    n = write_xlsx(form, out_path)
    print(f"wrote {n} rows -> {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
