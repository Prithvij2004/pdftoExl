from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.models import ExtractedRow, QuestionType
from app.services.excel_writer import HEADERS, write_rows_to_xlsx


def test_write_rows_to_xlsx_includes_section_column(tmp_path: Path) -> None:
    out_path = tmp_path / "out.xlsx"
    rows = [
        ExtractedRow(
            section="Member Information",
            sequence=1,
            question_type=QuestionType.TEXT_BOX,
            question_text="Full name",
            answer_text="",
            page_number=1,
            source_order=0,
        )
    ]

    write_rows_to_xlsx(rows, out_path)

    wb = load_workbook(out_path)
    ws = wb.active

    assert [ws.cell(row=1, column=i).value for i in range(1, 6)] == HEADERS
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Member Information"
    assert ws.cell(row=2, column=3).value == "Text Box"
    assert ws.cell(row=2, column=4).value == "Full name"
    assert ws.cell(row=2, column=5).value is None
    assert ws.auto_filter.ref == "A1:E1"

    wb.close()
