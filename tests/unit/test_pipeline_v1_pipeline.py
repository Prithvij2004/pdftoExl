"""Tests for PipelineV1 orchestration: stage toggles and build_pipeline."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pdftoxl.pipeline_v1 import PipelineConfig
from pdftoxl.pipeline_v1.pipeline import PipelineContext, PipelineV1, build_pipeline


def _template(path: Path, sheet: str, header_row: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for col in range(1, 4):
        ws.cell(row=header_row, column=col, value=f"H{col}")
    wb.save(path)


@pytest.fixture
def ctx(tmp_path):
    template = tmp_path / "tpl.xlsx"
    _template(template, "Assessment", 2)
    return PipelineContext(
        out_path=tmp_path / "out.xlsx",
        template_xlsx=template,
        question_sheet="Assessment",
        header_row=2,
    )


def test_pipeline_end_to_end_writes_output(tmp_path, ctx):
    pdf = tmp_path / "in.pdf"
    pdf.write_bytes(b"%PDF-1.4")
    pipeline = PipelineV1(PipelineConfig(), ctx)
    out = pipeline(pdf)
    assert out == ctx.out_path
    assert load_workbook(out).sheetnames == ["Assessment"]


def test_pipeline_with_all_stages_disabled_still_writes(tmp_path, ctx):
    """Every stage toggle off must not crash — output stage always runs regardless."""
    pdf = tmp_path / "in.pdf"
    pdf.write_bytes(b"%PDF-1.4")
    cfg = PipelineConfig()
    for name in ("extraction", "classification", "gate", "llm", "merge", "mapping"):
        setattr(cfg.stages, name, False)
    pipeline = PipelineV1(cfg, ctx)
    out = pipeline(pdf)
    assert out.exists()


def test_build_pipeline_does_not_construct_bedrock_without_with_llm(ctx, monkeypatch):
    """Default: no Bedrock client built — so no AWS deps required."""
    called = {"n": 0}

    def _boom(*a, **kw):
        called["n"] += 1
        raise AssertionError("bedrock client should not be built")

    monkeypatch.setattr(
        "pdftoxl.pipeline_v1.pipeline.build_bedrock_client", _boom
    )
    pipeline = build_pipeline(PipelineConfig(), ctx, with_llm=False)
    assert isinstance(pipeline, PipelineV1)
    assert pipeline._llm_client is None
    assert called["n"] == 0


def test_build_pipeline_constructs_client_when_with_llm_true(ctx, monkeypatch):
    sentinel = object()
    captured = {}

    def _fake_build(settings):
        captured["settings"] = settings
        return sentinel

    monkeypatch.setattr(
        "pdftoxl.pipeline_v1.pipeline.build_bedrock_client", _fake_build
    )
    cfg = PipelineConfig()
    pipeline = build_pipeline(cfg, ctx, with_llm=True)
    assert pipeline._llm_client is sentinel
    assert captured["settings"].model_id == cfg.bedrock.model_id
    assert captured["settings"].region == cfg.bedrock.region


def test_build_pipeline_skips_client_when_llm_stage_disabled(ctx, monkeypatch):
    """with_llm=True but stages.llm=False → still no client (avoids wasted AWS init)."""
    monkeypatch.setattr(
        "pdftoxl.pipeline_v1.pipeline.build_bedrock_client",
        lambda s: (_ for _ in ()).throw(AssertionError("should not build")),
    )
    cfg = PipelineConfig()
    cfg.stages.llm = False
    pipeline = build_pipeline(cfg, ctx, with_llm=True)
    assert pipeline._llm_client is None
