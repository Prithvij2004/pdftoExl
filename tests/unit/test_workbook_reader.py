from pathlib import Path

from openpyxl import Workbook

from pdftoxl.evals.workbook import read_sheet, read_workbook


def _write_tmp_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment v2"
    ws["A1"] = "header meta"
    ws["A3"] = "Section"
    ws["B3"] = "Sequence"
    ws["A4"] = "Intake"
    ws["B4"] = 1
    ws["A5"] = "Intake"
    ws["B5"] = 2
    ws.merge_cells("A1:B1")
    ws.freeze_panes = "A4"
    vs = wb.create_sheet("Values")
    vs["C1"] = "QuestionType"
    vs["C2"] = "Text Box"
    vs["C3"] = "Radio Button"
    path = tmp_path / "x.xlsx"
    wb.save(path)
    return path


def test_read_sheet_basic(tmp_path):
    path = _write_tmp_xlsx(tmp_path)
    s = read_sheet(path, "Assessment v2", 3)
    assert s.headers == ["Section", "Sequence"]
    assert s.row_count == 2
    assert s.cells[(4, 1)].value == "Intake"
    assert s.cells[(5, 2)].value == 2


def test_merged_cell_expanded(tmp_path):
    path = _write_tmp_xlsx(tmp_path)
    s = read_sheet(path, "Assessment v2", 3)
    assert s.cells[(1, 1)].value == "header meta"
    assert s.cells[(1, 2)].value == "header meta"
    assert s.cells[(1, 2)].merged_from == (1, 1)


def test_read_workbook_and_vocab(tmp_path):
    from pdftoxl.evals.workbook import read_values_vocabulary

    path = _write_tmp_xlsx(tmp_path)
    wb = read_workbook(path, "Assessment v2", 3)
    assert "Values" in wb.sheet_names
    vocab = read_values_vocabulary(path)
    assert "Text Box" in vocab
    assert "Radio Button" in vocab


def test_formula_vs_value(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessment v2"
    ws["A3"] = "Section"
    ws["B3"] = "Seq"
    ws["A4"] = "Intake"
    ws["B4"] = "=1+1"
    path = tmp_path / "f.xlsx"
    wb.save(path)
    s = read_sheet(path, "Assessment v2", 3)
    c = s.cells[(4, 2)]
    assert c.formula == "=1+1"
