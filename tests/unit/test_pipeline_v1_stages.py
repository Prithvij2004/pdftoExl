"""Unit tests for each PipelineV1 stage's `run` function."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from pdftoxl.pipeline_v1.stages import (
    classification,
    extraction,
    gate,
    llm,
    mapping,
    merge,
    output,
)


def test_extraction_returns_empty_blocks_for_any_path(tmp_path):
    pdf = tmp_path / "dummy.pdf"
    pdf.write_bytes(b"%PDF-1.4 stub")
    raw = extraction.run(pdf)
    assert raw.pdf_path == pdf
    assert raw.blocks == []


def test_classification_passes_blocks_through():
    raw = extraction.RawBlocks(pdf_path=Path("x"), blocks=[{"a": 1}, {"a": 2}])
    out = classification.run(raw)
    assert out.blocks == raw.blocks


def test_gate_splits_by_confidence_threshold():
    classified = classification.ClassifiedBlocks(
        blocks=[
            {"id": 1, "confidence": 0.9},
            {"id": 2, "confidence": 0.5},
            {"id": 3, "confidence": 0.75},  # boundary is accepted (>=)
            {"id": 4},  # missing confidence → 0.0 → deferred
        ]
    )
    out = gate.run(classified, confidence_threshold=0.75)
    accepted_ids = {b["id"] for b in out.accepted}
    deferred_ids = {b["id"] for b in out.deferred}
    assert accepted_ids == {1, 3}
    assert deferred_ids == {2, 4}


def test_gate_handles_empty_input():
    out = gate.run(classification.ClassifiedBlocks(blocks=[]), confidence_threshold=0.5)
    assert out.accepted == [] and out.deferred == []


def test_llm_skips_when_no_client():
    gated = gate.GateOutput(accepted=[], deferred=[{"id": 1}])
    out = llm.run(gated, client=None)
    assert out.enriched == gated.deferred


def test_llm_skips_when_no_deferred_even_with_client():
    class _FakeClient:
        def invoke(self, prompt, *, system=None):
            raise AssertionError("should not be called")

    gated = gate.GateOutput(accepted=[{"id": 1}], deferred=[])
    out = llm.run(gated, client=_FakeClient())
    assert out.enriched == []


def test_llm_with_client_and_deferred_returns_blocks_unchanged():
    """The scaffold does not mutate — it just round-trips deferred blocks."""

    class _FakeClient:
        def invoke(self, prompt, *, system=None):
            return "ignored"

    deferred = [{"id": 1, "confidence": 0.1}]
    gated = gate.GateOutput(accepted=[], deferred=deferred)
    out = llm.run(gated, client=_FakeClient())
    assert out.enriched == deferred


def test_merge_combines_and_filters_by_min_confidence():
    gated = gate.GateOutput(
        accepted=[{"id": 1, "confidence": 0.9}, {"id": 2}],  # id=2 missing → defaults to 1.0
        deferred=[],
    )
    llm_out = llm.LLMOutput(
        enriched=[{"id": 3, "confidence": 0.04}, {"id": 4, "confidence": 0.5}]
    )
    merged = merge.run(gated, llm_out, min_confidence=0.05)
    ids = [b["id"] for b in merged.blocks]
    # id=3 drops (below threshold); id=2 survives because default is 1.0.
    assert ids == [1, 2, 4]


def test_mapping_returns_empty_scaffold():
    merged = merge.MergedBlocks(blocks=[{"id": 1}])
    mapped = mapping.run(merged)
    assert mapped.rows == []


def _fake_template(path: Path, sheet: str, header_row: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for col in range(1, 5):
        ws.cell(row=header_row, column=col, value=f"H{col}")
    # Dummy data rows that should be cleared on output.
    ws.cell(row=header_row + 1, column=1, value="existing-1")
    ws.cell(row=header_row + 2, column=2, value="existing-2")
    wb.save(path)


def test_output_clones_template_and_clears_rows_below_header(tmp_path):
    template = tmp_path / "template.xlsx"
    _fake_template(template, sheet="Assessment", header_row=3)
    out_path = tmp_path / "nested" / "out.xlsx"
    written = output.run(
        mapping.MappedRows(rows=[]),
        out_path=out_path,
        template_xlsx=template,
        question_sheet="Assessment",
        header_row=3,
    )
    assert written == out_path
    assert out_path.exists()
    wb = load_workbook(out_path)
    ws = wb["Assessment"]
    # Header preserved.
    assert ws.cell(row=3, column=1).value == "H1"
    # Rows below header cleared.
    assert ws.cell(row=4, column=1).value is None
    assert ws.cell(row=5, column=2).value is None


def test_output_tolerates_missing_sheet(tmp_path):
    """If the question sheet is absent, output should still produce a valid file."""
    template = tmp_path / "template.xlsx"
    _fake_template(template, sheet="Other", header_row=2)
    out_path = tmp_path / "out.xlsx"
    written = output.run(
        mapping.MappedRows(rows=[]),
        out_path=out_path,
        template_xlsx=template,
        question_sheet="NotThere",
        header_row=2,
    )
    assert written.exists()
    wb = load_workbook(written)
    assert "Other" in wb.sheetnames
