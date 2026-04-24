from pathlib import Path

from openpyxl import load_workbook

from pdftoxl.evals.fixtures import find_fixture, load_fixtures
from pdftoxl.evals.metrics.eval_c import run_eval_c
from pdftoxl.pipeline_v1 import PipelineConfig
from pdftoxl.pipeline_v1.pipeline import PipelineContext, build_pipeline

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURES_YAML = REPO_ROOT / "evals" / "fixtures.yaml"


def test_pipeline_v1_writes_openable_workbook(tmp_path):
    fx = find_fixture(load_fixtures(FIXTURES_YAML), "FX-TXLTSS-001")
    out = tmp_path / "out.xlsx"
    cfg = PipelineConfig()
    ctx = PipelineContext(
        out_path=out,
        template_xlsx=fx.golden_xlsx_path,
        question_sheet=fx.question_sheet,
        header_row=fx.header_row,
    )
    pipeline = build_pipeline(cfg, ctx, with_llm=False)
    written = pipeline(fx.pdf_path)
    assert written == out
    assert out.exists()
    wb = load_workbook(out)
    assert fx.question_sheet in wb.sheetnames


def test_pipeline_v1_satisfies_protocol(tmp_path):
    """PipelineV1 is accepted by the eval runner that expects the Pipeline protocol."""
    fx = find_fixture(load_fixtures(FIXTURES_YAML), "FX-TXLTSS-001")
    out = tmp_path / "out.xlsx"
    ctx = PipelineContext(
        out_path=out,
        template_xlsx=fx.golden_xlsx_path,
        question_sheet=fx.question_sheet,
        header_row=fx.header_row,
    )
    pipeline = build_pipeline(PipelineConfig(), ctx, with_llm=False)
    # Must not raise — structural typing only.
    result = run_eval_c(fx, pipeline)
    assert result.fixture_id == fx.id
