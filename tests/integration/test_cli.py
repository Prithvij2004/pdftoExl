"""Integration tests for the `pdftoxl` CLI."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from pdftoxl.cli import app

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURES_YAML = REPO_ROOT / "evals" / "fixtures.yaml"

runner = CliRunner()


def test_cli_run_writes_workbook(tmp_path):
    out = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "run",
            "--fixture", "FX-TXLTSS-001",
            "--out", str(out),
            "--fixtures-yaml", str(FIXTURES_YAML),
        ],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()
    # CLI prints the written path on success.
    assert str(out) in result.stdout
    wb = load_workbook(out)
    assert "Assessment" in wb.sheetnames


def test_cli_run_accepts_custom_config(tmp_path):
    cfg = tmp_path / "cfg.yaml"
    cfg.write_text(
        "stages:\n"
        "  llm: false\n"
        "logging:\n"
        "  level: WARNING\n"
    )
    out = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "run",
            "--fixture", "FX-TXLTSS-001",
            "--out", str(out),
            "--fixtures-yaml", str(FIXTURES_YAML),
            "--config", str(cfg),
        ],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()


def test_cli_missing_fixture_fails_nonzero(tmp_path):
    result = runner.invoke(
        app,
        [
            "run",
            "--fixture", "FX-DOES-NOT-EXIST",
            "--out", str(tmp_path / "out.xlsx"),
            "--fixtures-yaml", str(FIXTURES_YAML),
        ],
    )
    assert result.exit_code != 0


def test_cli_no_args_shows_help():
    result = runner.invoke(app, [])
    assert result.exit_code != 0 or "Usage" in result.output
