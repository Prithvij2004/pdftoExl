from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import ExtractedRow


HEADERS = [
    "Sequence",
    "Section",
    "Question Type",
    "English Question/Index Text",
    "English Answer Text",
]


_BOLD_RE = re.compile(r"\*\*(.+?)\*\*", re.DOTALL)
_BOLD_FONT = InlineFont(b=True)


def _to_rich_text(text: str) -> str | CellRichText:
    """Convert `**bold**` Markdown spans into an openpyxl rich-text cell value.

    Plain text (no bold markers) is returned unchanged so existing behavior is preserved.
    """
    if not text or "**" not in text:
        return text

    parts: list[object] = []
    last = 0
    for m in _BOLD_RE.finditer(text):
        if m.start() > last:
            parts.append(text[last:m.start()])
        parts.append(TextBlock(_BOLD_FONT, m.group(1)))
        last = m.end()
    if last < len(text):
        parts.append(text[last:])

    if not any(isinstance(p, TextBlock) for p in parts):
        return text.replace("**", "")

    return CellRichText(parts)


def write_rows_to_xlsx(rows: list[ExtractedRow], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E1"

    body_alignment = Alignment(vertical="top", wrap_text=True)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.sequence or i - 1).alignment = body_alignment
        ws.cell(row=i, column=2, value=r.section).alignment = body_alignment
        ws.cell(row=i, column=3, value=r.question_type.value).alignment = body_alignment
        ws.cell(row=i, column=4, value=_to_rich_text(r.question_text)).alignment = body_alignment
        ws.cell(row=i, column=5, value=r.answer_text).alignment = body_alignment

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 55

    # Row height (let Excel auto-fit visually; keep a minimum height for header)
    ws.row_dimensions[1].height = 24

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path

