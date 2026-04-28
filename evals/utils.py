from __future__ import annotations

import os
import re
import time
from dataclasses import dataclass
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


DOCS_DIR = Path(__file__).resolve().parents[1] / "docs"

# Hard-coded fixture pairing (because docs is gitignored and filenames are not guaranteed
# to be derived mechanically).
DEFAULT_CASES = [
    (
        "CHOICES Safety Determination Form.pdf",
        "CHOICES Safety Determination Request Form Final_11_20.xlsx",
    ),
    (
        "sph_rev25-3_H1700-3_final_approved.pdf",
        "TX LTSS - 1700-3, Individual Service Plan - Signature Page.xlsx",
    ),
]

EXPECTED_HEADERS = (
    "Sequence",
    "Section",
    "Question Type",
    "English Question/Index Text",
    "English Answer Text",
)
CORE_HEADERS = ("Question Type", "English Question/Index Text", "English Answer Text")

_WS_RE = re.compile(r"[ \t]+")


def _clean_cell_text(s: object) -> str:
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = s.replace("\u00a0", " ").strip()
    s = _WS_RE.sub(" ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s


def docs_available() -> bool:
    return DOCS_DIR.exists() and DOCS_DIR.is_dir()


@dataclass(frozen=True)
class EvalCase:
    case_id: str
    pdf_path: Path
    expected_xlsx_path: Path


def iter_eval_cases(
    docs_dir: Path = DOCS_DIR,
    cases: Iterable[tuple[str, str]] = DEFAULT_CASES,
) -> list[EvalCase]:
    out: list[EvalCase] = []
    for pdf_name, xlsx_name in cases:
        pdf_path = docs_dir / pdf_name
        expected_path = docs_dir / xlsx_name
        case_id = Path(pdf_name).stem
        out.append(EvalCase(case_id=case_id, pdf_path=pdf_path, expected_xlsx_path=expected_path))
    return out


def _core_column_indices(header_values: list[str]) -> tuple[int, int, int]:
    return tuple(header_values.index(name) + 1 for name in CORE_HEADERS)  # type: ignore[return-value]


def load_xlsx_rows_from_path(xlsx_path: Path) -> tuple[tuple[str, ...], list[tuple[str, str, str]]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    header = tuple(_clean_cell_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1))
    qt_col, q_col, a_col = _core_column_indices(list(header))

    rows: list[tuple[str, str, str]] = []
    for r in range(2, ws.max_row + 1):
        qt = _clean_cell_text(ws.cell(row=r, column=qt_col).value)
        q = _clean_cell_text(ws.cell(row=r, column=q_col).value)
        a = _clean_cell_text(ws.cell(row=r, column=a_col).value)
        if not (qt or q or a):
            continue
        if not q:
            continue
        rows.append((qt, q, a))
    wb.close()
    return header, rows


def load_xlsx_rows_from_bytes(xlsx_bytes: bytes) -> tuple[tuple[str, ...], list[tuple[str, str, str]]]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    header = tuple(_clean_cell_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1))
    qt_col, q_col, a_col = _core_column_indices(list(header))
    rows: list[tuple[str, str, str]] = []
    for r in range(2, ws.max_row + 1):
        qt = _clean_cell_text(ws.cell(row=r, column=qt_col).value)
        q = _clean_cell_text(ws.cell(row=r, column=q_col).value)
        a = _clean_cell_text(ws.cell(row=r, column=a_col).value)
        if not (qt or q or a):
            continue
        if not q:
            continue
        rows.append((qt, q, a))
    wb.close()
    return header, rows


class UnsupportedGoldenWorkbook(ValueError):
    pass


def load_golden_rows_from_path(
    xlsx_path: Path,
    *,
    header_row_scan: int = 250,
    max_rows: int = 10000,
    max_cols: int = 80,
) -> list[tuple[str, str, str]]:
    """
    Loads the *expected* rows from a golden workbook.

    Supports two golden formats:
    - A simple Extracted sheet (the same format the API returns).
    - A template-style sheet that includes columns named:
        Question Type, English Question/Index Text, English Answer Text
      in some header row; this function finds that row and extracts those columns.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)

    # 1) If any sheet is already in the 3-column format, use it.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [_clean_cell_text(ws.cell(row=1, column=c).value) for c in range(1, min(ws.max_column, max_cols) + 1)]
        if all(name in header for name in CORE_HEADERS):
            qt_col, q_col, a_col = _core_column_indices(header)
            rows: list[tuple[str, str, str]] = []
            for r in range(2, min(ws.max_row, max_rows) + 1):
                qt = _clean_cell_text(ws.cell(row=r, column=qt_col).value)
                q = _clean_cell_text(ws.cell(row=r, column=q_col).value)
                a = _clean_cell_text(ws.cell(row=r, column=a_col).value)
                if not (qt or q or a):
                    continue
                if not q:
                    continue
                rows.append((qt, q, a))
            wb.close()
            return rows

    # 2) Otherwise, find a header row in any sheet with the required column names.
    header_targets = CORE_HEADERS
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(header_row_scan, ws.max_row), max_col=max_cols, values_only=True),
            start=1,
        ):
            if not row:
                continue
            vals = [_clean_cell_text(v) for v in row]
            if not vals:
                continue
            if all(t in vals for t in header_targets):
                qt_col = vals.index(header_targets[0]) + 1
                q_col = vals.index(header_targets[1]) + 1
                a_col = vals.index(header_targets[2]) + 1

                out: list[tuple[str, str, str]] = []
                for rr in range(r_idx + 1, min(ws.max_row, max_rows) + 1):
                    qt = _clean_cell_text(ws.cell(row=rr, column=qt_col).value)
                    q = _clean_cell_text(ws.cell(row=rr, column=q_col).value)
                    a = _clean_cell_text(ws.cell(row=rr, column=a_col).value)
                    if not (qt or q or a):
                        continue
                    if not q:
                        continue
                    out.append((qt, q, a))
                wb.close()
                if not out:
                    raise UnsupportedGoldenWorkbook(f"{xlsx_path} contains headers but no usable rows in sheet {sheet_name}")
                return out

    wb.close()
    raise UnsupportedGoldenWorkbook(
        f"{xlsx_path} does not contain a usable golden row table with columns: {CORE_HEADERS}"
    )


def _sim(a: str, b: str) -> float:
    a = (a or "").strip()
    b = (b or "").strip()
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


@dataclass
class EvalScore:
    expected_rows: int
    actual_rows: int
    matched_rows: int
    coverage: float
    type_accuracy: float
    avg_answer_similarity: float
    extra_rows: int
    missing_rows: int
    elapsed_s: Optional[float] = None

    def rows_per_second(self) -> Optional[float]:
        if self.elapsed_s is None or self.elapsed_s <= 0:
            return None
        return self.actual_rows / self.elapsed_s


@dataclass(frozen=True)
class MatchDetail:
    expected: tuple[str, str, str]
    actual: Optional[tuple[str, str, str]]
    question_similarity: float
    answer_similarity: float


def score_rows(
    expected: list[tuple[str, str, str]],
    actual: list[tuple[str, str, str]],
    *,
    min_question_similarity: float = 0.72,
    answer_weight: float = 0.25,
    top_k_scan: int = 200,
) -> tuple[EvalScore, list[MatchDetail], list[tuple[str, str, str]]]:
    """
    Greedy matching: for each expected row, find best unmatched actual row by a weighted
    similarity of question + answer text. This is robust to reordering and small wording
    changes, but will expose missing/extra rows clearly.
    """
    expected_n = len(expected)
    actual_n = len(actual)

    unmatched_actual = list(actual)
    matches: list[MatchDetail] = []

    matched = 0
    type_correct = 0
    answer_sims: list[float] = []

    for e_qt, e_q, e_a in expected:
        best_idx = None
        best_score = -1.0
        best_q_sim = 0.0
        best_a_sim = 0.0

        scan = unmatched_actual[:top_k_scan] if len(unmatched_actual) > top_k_scan else unmatched_actual
        for i, (a_qt, a_q, a_a) in enumerate(scan):
            q_sim = _sim(e_q, a_q)
            if q_sim < min_question_similarity:
                continue
            a_sim = _sim(e_a, a_a)
            score = (1.0 - answer_weight) * q_sim + answer_weight * a_sim
            if score > best_score:
                best_score = score
                best_idx = i
                best_q_sim = q_sim
                best_a_sim = a_sim

        if best_idx is None:
            matches.append(MatchDetail(expected=(e_qt, e_q, e_a), actual=None, question_similarity=0.0, answer_similarity=0.0))
            continue

        a_row = unmatched_actual.pop(best_idx)
        matched += 1
        if (a_row[0] or "").strip() == (e_qt or "").strip():
            type_correct += 1
        answer_sims.append(best_a_sim)
        matches.append(MatchDetail(expected=(e_qt, e_q, e_a), actual=a_row, question_similarity=best_q_sim, answer_similarity=best_a_sim))

    coverage = matched / expected_n if expected_n else 1.0
    type_acc = type_correct / matched if matched else 0.0
    avg_a_sim = sum(answer_sims) / len(answer_sims) if answer_sims else 0.0

    score = EvalScore(
        expected_rows=expected_n,
        actual_rows=actual_n,
        matched_rows=matched,
        coverage=coverage,
        type_accuracy=type_acc,
        avg_answer_similarity=avg_a_sim,
        extra_rows=len(unmatched_actual),
        missing_rows=expected_n - matched,
    )
    return score, matches, unmatched_actual


def require_env_flag(name: str) -> bool:
    v = os.getenv(name, "").strip().lower()
    return v in {"1", "true", "yes", "y", "on"}


def now_s() -> float:
    return time.perf_counter()

