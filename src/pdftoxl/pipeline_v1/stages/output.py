"""Stage 7: write the mapped rows into a workbook.

Until mapping is implemented the output is a clone of the golden template
with the question rows cleared below the header. That keeps the workbook
valid/openable while producing an empty candidate the evals can score.
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from ...adapters.logging import get_logger
from .mapping import MappedRows

log = get_logger("output")


def run(
    mapped: MappedRows,
    *,
    out_path: Path,
    template_xlsx: Path,
    question_sheet: str,
    header_row: int,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_xlsx, out_path)

    wb = load_workbook(out_path)
    if question_sheet in wb.sheetnames:
        ws = wb[question_sheet]
        # Clear rows below the header so candidate starts empty.
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.value = None

    # `mapped.rows` is empty in the scaffold; a later issue fills the sheet.
    wb.save(out_path)
    log.info("output", out=str(out_path), rows=len(mapped.rows))
    return out_path
