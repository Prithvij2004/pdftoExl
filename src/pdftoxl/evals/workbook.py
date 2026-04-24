from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class Cell:
    row: int
    col: int
    value: Any
    formula: str | None = None
    is_merged_origin: bool = False
    merged_from: tuple[int, int] | None = None


@dataclass
class SheetSnapshot:
    name: str
    header_row: int
    headers: list[str]
    data_start_row: int
    data_end_row: int
    cells: dict[tuple[int, int], Cell] = field(default_factory=dict)
    merged_ranges: list[str] = field(default_factory=list)
    frozen_panes: str | None = None
    data_validations: list[str] = field(default_factory=list)

    def column_index(self, header: str) -> int | None:
        for i, h in enumerate(self.headers, start=1):
            if (h or "").strip() == header.strip():
                return i
        return None

    def column_values(self, col: int) -> list[Any]:
        values = []
        for r in range(self.data_start_row, self.data_end_row + 1):
            cell = self.cells.get((r, col))
            values.append(cell.value if cell else None)
        return values

    @property
    def row_count(self) -> int:
        return max(0, self.data_end_row - self.data_start_row + 1)


@dataclass
class WorkbookSnapshot:
    path: Path
    sheet_names: list[str]
    named_ranges: list[str]
    question_sheet: SheetSnapshot
    defined_names: list[str] = field(default_factory=list)


def _load(path: Path, data_only: bool) -> Workbook:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(path, data_only=data_only)


def _last_non_empty_row(ws: Worksheet, header_row: int, max_col: int) -> int:
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                last = r
                break
    return last


def _expand_merges(ws: Worksheet, cells: dict[tuple[int, int], Cell]) -> list[str]:
    ranges = []
    for mr in ws.merged_cells.ranges:
        ranges.append(str(mr))
        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        origin = cells.get((min_row, min_col))
        if origin is None:
            continue
        origin.is_merged_origin = True
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) == (min_row, min_col):
                    continue
                cells[(r, c)] = Cell(
                    row=r,
                    col=c,
                    value=origin.value,
                    formula=origin.formula,
                    merged_from=(min_row, min_col),
                )
    return ranges


def read_sheet(
    path: Path,
    sheet_name: str,
    header_row: int,
) -> SheetSnapshot:
    path = Path(path)
    wb_values = _load(path, data_only=True)
    wb_formulas = _load(path, data_only=False)
    if sheet_name not in wb_values.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name} in {path}")
    ws_v = wb_values[sheet_name]
    ws_f = wb_formulas[sheet_name]

    max_col = ws_v.max_column or 1
    headers: list[str] = []
    for c in range(1, max_col + 1):
        v = ws_v.cell(header_row, c).value
        headers.append("" if v is None else str(v))
    while headers and (headers[-1] or "").strip() == "":
        headers.pop()
    max_col = len(headers)

    cells: dict[tuple[int, int], Cell] = {}
    for r in range(1, ws_v.max_row + 1):
        for c in range(1, max_col + 1):
            vv = ws_v.cell(r, c).value
            fv = ws_f.cell(r, c).value
            formula = None
            if isinstance(fv, str) and fv.startswith("="):
                formula = fv
            if vv is None and formula is None and fv is None:
                continue
            cells[(r, c)] = Cell(row=r, col=c, value=vv, formula=formula)

    merged_ranges = _expand_merges(ws_v, cells)

    last_row = _last_non_empty_row(ws_v, header_row, max_col)
    data_start = header_row + 1
    data_end = max(last_row, header_row)

    frozen = ws_v.freeze_panes
    validations = []
    try:
        for dv in ws_v.data_validations.dataValidation:
            for rng in dv.sqref.ranges:
                validations.append(str(rng))
    except Exception:
        pass

    return SheetSnapshot(
        name=sheet_name,
        header_row=header_row,
        headers=headers,
        data_start_row=data_start,
        data_end_row=data_end,
        cells=cells,
        merged_ranges=merged_ranges,
        frozen_panes=frozen,
        data_validations=validations,
    )


def read_workbook(path: Path, question_sheet: str, header_row: int) -> WorkbookSnapshot:
    path = Path(path)
    wb = _load(path, data_only=True)
    snapshot = read_sheet(path, question_sheet, header_row)
    defined = [n for n in wb.defined_names]
    return WorkbookSnapshot(
        path=path,
        sheet_names=list(wb.sheetnames),
        named_ranges=defined,
        question_sheet=snapshot,
        defined_names=defined,
    )


def read_values_vocabulary(path: Path, column_letter: str = "C") -> list[str]:
    wb = _load(Path(path), data_only=True)
    if "Values" not in wb.sheetnames:
        return []
    ws = wb["Values"]
    col_idx = openpyxl.utils.column_index_from_string(column_letter)
    vocab: list[str] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() != "questiontype" and s.lower() != "question type":
            vocab.append(s)
    # Header row values are sometimes present; filter but keep uniqueness
    seen = set()
    out = []
    for v in vocab:
        k = v.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(v)
    return out


def coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"
